
import pandas as pd
import streamlit as st
from dataclasses import dataclass,field
from io import BytesIO
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup

@dataclass
class Budget:
    year : int
    receipts: int
    outlays : int
    deficit : int 

@dataclass
class Expenditure:
    name : str
    amount : int
    percentage : float
    level : int
    gao_function: GAOFunction | None = None

@dataclass
class GAOFunction:
    code: str
    name: str
    description: str
    supercode: str | None = None
    subcodes: list[str]=field(default_factory=list)

def load_css_for_streamlit_controls(filename):
    with open(filename) as f:
        st.html(f"<style>{f.read()}</style>")

def get_budget_data()-> Budget:
    dataset_url = "https://www.govinfo.gov/content/pkg/BUDGET-2027-TAB/xls/BUDGET-2027-TAB-2-1.xlsx"
    df = pd.read_excel(dataset_url)
    latest = df.dropna(subset=[df.columns[1]]).iloc[-1]
    return Budget(
        year=int(latest.iloc[0]),
        receipts=int(latest.iloc[1]),
        outlays=int(latest.iloc[2]),
        deficit=int(latest.iloc[3])
    )

def get_expenditure_by_function(budget_total: int,year:int = 2025)-> list[Expenditure]:
    gao_by_name = {
        x.name.strip().lower(): x
        for x in get_GAO_functions()
    }
    dataset_url = (
        "https://www.govinfo.gov/content/pkg/"
        "BUDGET-2027-TAB/xls/BUDGET-2027-TAB-4-1.xlsx"
        )
    response = requests.get(dataset_url,timeout=30)
    wb=load_workbook(BytesIO(response.content))
    ws=wb.active
    
    year_column = None
    for cell in ws[2]:
        if cell.value == str(year):
            year_column = cell.column
            break
    if year_column is None:
        raise ValueError(f"Year {year} not found in expenditure table")

    expenditures = []

    for row in range(4, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1)

        if name_cell.value == "As percentages of outlays:":
            break

        amount_cell = ws.cell(row=row, column=year_column)

        name = name_cell.value
        amount = amount_cell.value

        if name is None:
            continue

        if amount is None or str(amount).strip(". ") == "":
            amount = 0

        expenditures.append(
            Expenditure(
                name=str(name),
                amount=int(amount),
                percentage=(int(amount) / budget_total) * 100,
                level=int(name_cell.alignment.indent or 0),
                gao_function=gao_by_name.get(name.lower())
            )
        )
    
    return expenditures

def get_major_categories(expenditures:list[Expenditure])-> list[Expenditure]:
    top_level_functions = [
        x for x in expenditures
        if x.level==1 
        or x.name == "National Defense" 
        or x.name == "Net interest"
        or (x.name == "Undistributed offsetting receipts" and x.level==0)
    ]

    functions = sorted(
        top_level_functions,
        key= lambda x: x.amount,
        reverse=True)

    major_categories = [
        x
        for x in functions
        if x.percentage >= 10 
    ]

    categories_in_others = [
        x for x in functions
        if x.percentage<10
        ]
    
    other_amount = sum(x.amount 
                       for x in categories_in_others
                       if x.amount>0)
    other_percentage = sum(x.percentage 
                           for x in categories_in_others
                           if x.amount>0)

    other_expenditure = Expenditure(
        name="Others",
        amount=other_amount,
        percentage=other_percentage,
        level=0
    )

    uor_amount = sum(x.amount
                             for x in categories_in_others
                             if x.amount<0)

    uor_percentage = sum(x.percentage 
                                 for x in categories_in_others 
                                 if x.amount<0)

    uor_expenditure = Expenditure(
        name="UOR",
        amount=uor_amount,
        percentage=uor_percentage,
        level=0
    )
    
    major_categories.append(other_expenditure)
    major_categories.append(uor_expenditure)

    return major_categories

def get_GAO_functions()->list[GAOFunction]: 
    """ This function parses the GAO glossary expecting the following format:
    SuperCode
    SuperFunction
    SuperDescription
        Subcode 1
        Subfunction 1
        Subdescription 1
        Subcode n
        Subfunction n
        Subdescription n
    SuperCode 2
    ...
    Source: GAO
    """
    with open("GAO.html","r", encoding="utf-8") as f:
        soup = BeautifulSoup(f,"html.parser")
    gao_text = soup.get_text("\n",strip=True)

    start_marker = "Code: 050;"
    end_marker = "Source: GAO."

    start_index = gao_text.find(start_marker)
    end_index = gao_text.find(end_marker)

    if start_index==-1 or end_index==-1:
        raise ValueError("Cannot locate definitions section")

    important_bits = gao_text[start_index:end_index]

    lines = important_bits.splitlines()

    supercode_marker = "Code:"
    subcode_marker = "Subcode:"
    function_marker = "Function:"
    subfunction_marker="Subfunction:"

    GAOFunctions=[]

    #Superfunction state
    current_supercode=None
    current_function=None
    current_superfunction_description=[]
    current_subcodes=[]

    #Subfunction state
    current_subcode=None
    current_subfunction=None
    current_subfunction_description=[]

    #When we hit a subcode, the description lines belong to the subcode
    subcode_marker_hit = False

    for line in lines:    

        #New Superfunction
        if supercode_marker in line:
            #If the previous superfunction had a final subfunction, the subfunction is now saved
            if current_subcode:
                            GAOFunctions.append(
                                GAOFunction(
                                    code=current_subcode,
                                    name=current_subfunction,
                                    description="\n".join(current_subfunction_description).strip(),
                                    supercode=current_supercode
                                )
                            )
            #The superfunction we're building is now also finished
            if current_supercode:
                GAOFunctions.append(
                    GAOFunction(
                        code=current_supercode,
                        name=current_function,
                        description="\n".join(current_superfunction_description).strip(),
                        subcodes=current_subcodes.copy()))
                
            #start a new superfunction
            current_supercode = line.removeprefix(supercode_marker).strip().removesuffix(";")
            current_function = None
            current_superfunction_description = []
            current_subcodes = []

            #clear subcode state, because there's a new superfunction
            current_subcode = None
            current_subfunction = None
            current_subfunction_description=[]
            subcode_marker_hit = False

        #Superfunction name
        elif function_marker in line:
            current_function = line.removeprefix(function_marker).strip().removesuffix(".")

        #new Subfunction
        elif subcode_marker in line:
            #if previous subfunction exists, save it now
            if current_subcode:
                GAOFunctions.append(
                    GAOFunction(
                        code=current_subcode,
                        name=current_subfunction,
                        description="\n".join(current_subfunction_description).strip(),
                        supercode=current_supercode
                    )
                )
            #start new subfunction
            current_subcode= line.removeprefix(subcode_marker).strip().removesuffix(";")
            current_subfunction=None
            current_subfunction_description=[]

            subcode_marker_hit = True
            current_subcodes.append(current_subcode)
            
        elif subfunction_marker in line:
            current_subfunction= line.removeprefix(subfunction_marker).strip().removesuffix(".")
    
        else:
            if not subcode_marker_hit:
                current_superfunction_description.append(line)
            else:
                current_subfunction_description.append(line)

    #save last subfunction of last superfunction
    if current_subcode:
        GAOFunctions.append(
            GAOFunction(
                code=current_subcode,
                name=current_subfunction,
                description="\n".join(current_subfunction_description).strip(),
                supercode=current_supercode
            )
        )

    #save last superfunction
    if current_supercode:
        GAOFunctions.append(
            GAOFunction(
                code=current_supercode,
                name=current_function,
                description="\n".join(current_superfunction_description).strip(),
                subcodes=current_subcodes.copy()))
        
    return GAOFunctions


#%%
from etl import get_budget_data
from etl import get_expenditure_by_function
from etl import get_major_categories
from etl import get_GAO_functions
budget_total = get_budget_data().outlays
expenditures = get_expenditure_by_function(budget_total)
# major_categories = get_major_categories(expenditures)
expenditures[0]
# for x in major_categories:
#     print(f"{x.name:<55} : {x.percentage:>6.2f} %")
# gfs=get_GAO_functions()
# gfs[0]
# %%
