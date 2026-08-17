from etl.datamodel import Expenditure
from etl.get_GAO_functions import get_GAO_functions


import requests
import streamlit as st
from openpyxl import load_workbook


from io import BytesIO


@st.cache_data
def get_expenditure_by_function(budget_total: int,year:int = 2025)-> list[Expenditure]:
    gao_by_name = {
        x.name.strip().lower(): x
        for x in get_GAO_functions()
    }
    dataset_url = (
        "https://www.govinfo.gov/content/pkg/BUDGET-2027-TAB/xls/BUDGET-2027-TAB-4-1.xlsx"
        )
    response = requests.get(dataset_url,timeout=30)
    wb=load_workbook(BytesIO(response.content))
    ws=wb.active

    year_column = None
    for cell in ws[2]:
        if cell.value == str(year):
            year_column = cell.column
            break
    if year_column is None:
        raise ValueError(f"Year {year} not found in expenditure table")

    expenditures = []

    for row in range(4, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1)

        if name_cell.value == "As percentages of outlays:":
            break

        amount_cell = ws.cell(row=row, column=year_column)

        name = name_cell.value
        amount = amount_cell.value

        if name is None:
            continue

        if amount is None or str(amount).strip(". ") == "":
            amount = 0

        expenditures.append(
            Expenditure(
                name=str(name),
                amount=int(amount),
                percentage=(int(amount) / budget_total) * 100,
                level=int(name_cell.alignment.indent or 0),
                gao_function=gao_by_name.get(name.lower())
            )
        )

    return expenditures